"""
KPI Management UI for Managers - Manual Data Entry with Chart Generation
Professional dashboard for entering KPI data and visualizing results
"""

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from kpi_manager import KPIManager
from kpi_quarterly_calculator import KPIQuarterlyCalculator
from datetime import datetime
import traceback
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Matplotlib imports for chart generation
import matplotlib
matplotlib.use('Qt5Agg')
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt


class KPIDashboard(QWidget):
    """Professional KPI Dashboard with Manual Data Entry and Chart Generation"""

    def __init__(self, pool, current_user, parent=None):
        super().__init__(parent)
        self.pool = pool
        self.current_user = current_user
        self.kpi_manager = KPIManager(pool)
        self.quarterly_calculator = KPIQuarterlyCalculator(pool)
        self.current_period = datetime.now().strftime('%Y-%m')
        self.is_quarterly_period = False
        self.chart_canvas = None
        self.init_ui()

    def init_ui(self):
        """Initialize the professional user interface"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)

        # Header Section with gradient background
        header_widget = QWidget()
        header_widget.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #3498db);
                border-radius: 10px;
                padding: 20px;
            }
        """)
        header_layout = QVBoxLayout()

        title_label = QLabel("📊 KPI Dashboard 2025 - Professional Edition")
        title_label.setStyleSheet("font-size: 24pt; font-weight: bold; color: white;")
        title_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title_label)

        subtitle_label = QLabel("Manual Data Entry, Real-Time Visualization & Quarterly Reporting")
        subtitle_label.setStyleSheet("font-size: 12pt; color: #ecf0f1;")
        subtitle_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(subtitle_label)

        header_widget.setLayout(header_layout)
        main_layout.addWidget(header_widget)

        # Period and Controls Section
        controls_widget = QWidget()
        controls_widget.setStyleSheet("""
            QWidget {
                background-color: #f8f9fa;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        controls_layout = QHBoxLayout()

        period_label = QLabel("📅 Measurement Period:")
        period_label.setStyleSheet("font-size: 11pt; font-weight: bold; color: #2c3e50;")
        controls_layout.addWidget(period_label)

        self.period_combo = QComboBox()
        self.period_combo.setMinimumWidth(200)
        self.period_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #3498db;
                border-radius: 5px;
                font-size: 10pt;
                background-color: white;
            }
        """)
        self.populate_periods()
        self.period_combo.currentTextChanged.connect(self.on_period_changed)
        controls_layout.addWidget(self.period_combo)

        controls_layout.addStretch()

        # Quarterly Report Generation Button (only visible for quarterly periods)
        self.generate_quarterly_btn = QPushButton("📊 Generate & Save Quarterly Report")
        self.generate_quarterly_btn.setStyleSheet(self.get_button_style("#e67e22"))
        self.generate_quarterly_btn.clicked.connect(self.generate_quarterly_report)
        self.generate_quarterly_btn.setMinimumHeight(40)
        self.generate_quarterly_btn.setVisible(False)  # Hidden by default
        controls_layout.addWidget(self.generate_quarterly_btn)

        refresh_btn = QPushButton("🔄 Refresh Dashboard")
        refresh_btn.setStyleSheet(self.get_button_style("#95a5a6"))
        refresh_btn.clicked.connect(self.refresh_dashboard)
        refresh_btn.setMinimumHeight(40)
        controls_layout.addWidget(refresh_btn)

        controls_widget.setLayout(controls_layout)
        main_layout.addWidget(controls_widget)

        # Main Content Area with Tabs
        tab_widget = QTabWidget()
        tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background: #ecf0f1;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background: #3498db;
                color: white;
            }
        """)

        # Tab 1: Data Entry & Visualization
        self.entry_tab = self.create_data_entry_tab()
        tab_widget.addTab(self.entry_tab, "📝 Data Entry & Charts")

        # Tab 2: Overview Dashboard
        self.overview_tab = self.create_overview_tab()
        tab_widget.addTab(self.overview_tab, "📊 KPI Overview")

        # Tab 3: Export Reports
        self.export_tab = self.create_export_tab()
        tab_widget.addTab(self.export_tab, "📄 Export Reports")

        main_layout.addWidget(tab_widget)
        self.setLayout(main_layout)
        self.refresh_dashboard()

    def get_button_style(self, color):
        """Get professional button style"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-size: 10pt;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color, 0.7)};
            }}
        """

    def darken_color(self, hex_color, factor=0.8):
        """Darken a hex color"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        darker_rgb = tuple(int(c * factor) for c in rgb)
        return f"#{darker_rgb[0]:02x}{darker_rgb[1]:02x}{darker_rgb[2]:02x}"

    def create_data_entry_tab(self):
        """Create the main data entry and visualization tab"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)

        # KPI Selection Section
        selection_group = QGroupBox("📋 Select KPI for Data Entry")
        selection_group.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        selection_layout = QVBoxLayout()

        info_label = QLabel("Select a KPI below to enter data manually. After entering data, click 'Calculate & Generate Chart' to see results and visualizations.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e8f4f8; padding: 10px; border-radius: 5px; font-size: 10pt;")
        selection_layout.addWidget(info_label)

        kpi_select_layout = QHBoxLayout()
        kpi_select_layout.addWidget(QLabel("KPI:"))

        self.kpi_selector = QComboBox()
        self.kpi_selector.setMinimumHeight(40)
        self.kpi_selector.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #3498db;
                border-radius: 5px;
                font-size: 11pt;
                background-color: white;
            }
        """)
        self.populate_kpi_selector()
        self.kpi_selector.currentTextChanged.connect(self.on_kpi_selected)
        kpi_select_layout.addWidget(self.kpi_selector, 1)

        selection_layout.addLayout(kpi_select_layout)
        selection_group.setLayout(selection_layout)
        layout.addWidget(selection_group)

        # Splitter for Data Entry and Chart
        splitter = QSplitter(Qt.Horizontal)

        # Left side: Data Entry Form
        entry_widget = QWidget()
        entry_layout = QVBoxLayout()

        # KPI Info Display
        self.kpi_info_group = QGroupBox("ℹ️ KPI Information")
        self.kpi_info_group.setStyleSheet("""
            QGroupBox {
                font-size: 11pt;
                font-weight: bold;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding: 10px;
                background-color: #f1f9f5;
            }
        """)
        self.kpi_info_layout = QVBoxLayout()
        self.kpi_info_group.setLayout(self.kpi_info_layout)
        entry_layout.addWidget(self.kpi_info_group)

        # Data Input Form
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; }")

        self.input_form_widget = QWidget()
        self.input_form_widget.setStyleSheet("background-color: white;")
        self.input_form_layout = QVBoxLayout()
        self.input_form_widget.setLayout(self.input_form_layout)
        scroll.setWidget(self.input_form_widget)
        entry_layout.addWidget(scroll)

        # Action Buttons
        button_layout = QHBoxLayout()

        save_btn = QPushButton("💾 Save Data")
        save_btn.setStyleSheet(self.get_button_style("#95a5a6"))
        save_btn.clicked.connect(self.save_manual_data)
        save_btn.setMinimumHeight(50)
        button_layout.addWidget(save_btn)

        calc_chart_btn = QPushButton("📈 Calculate & Generate Chart")
        calc_chart_btn.setStyleSheet(self.get_button_style("#27ae60"))
        calc_chart_btn.clicked.connect(self.calculate_and_chart)
        calc_chart_btn.setMinimumHeight(50)
        button_layout.addWidget(calc_chart_btn)

        entry_layout.addLayout(button_layout)
        entry_widget.setLayout(entry_layout)

        # Right side: Chart Display
        chart_widget = QWidget()
        chart_layout = QVBoxLayout()

        chart_title = QLabel("📊 KPI Visualization")
        chart_title.setStyleSheet("font-size: 14pt; font-weight: bold; color: #2c3e50; padding: 10px;")
        chart_title.setAlignment(Qt.AlignCenter)
        chart_layout.addWidget(chart_title)

        self.chart_container = QWidget()
        self.chart_container.setMinimumSize(400, 400)
        self.chart_container.setStyleSheet("background-color: white; border: 2px solid #bdc3c7; border-radius: 8px;")
        self.chart_container_layout = QVBoxLayout()
        self.chart_container.setLayout(self.chart_container_layout)
        chart_layout.addWidget(self.chart_container)

        chart_widget.setLayout(chart_layout)

        splitter.addWidget(entry_widget)
        splitter.addWidget(chart_widget)
        splitter.setSizes([400, 500])

        layout.addWidget(splitter)
        widget.setLayout(layout)
        return widget

    def create_overview_tab(self):
        """Create KPI overview tab"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Summary cards
        summary_layout = QHBoxLayout()

        self.total_kpis_label = QLabel("Total KPIs\n0/17")
        self.total_kpis_label.setStyleSheet(self.get_card_style("#3498db"))
        summary_layout.addWidget(self.total_kpis_label)

        self.passing_kpis_label = QLabel("✓ Passing\n0")
        self.passing_kpis_label.setStyleSheet(self.get_card_style("#27ae60"))
        summary_layout.addWidget(self.passing_kpis_label)

        self.failing_kpis_label = QLabel("✗ Failing\n0")
        self.failing_kpis_label.setStyleSheet(self.get_card_style("#e74c3c"))
        summary_layout.addWidget(self.failing_kpis_label)

        self.pending_kpis_label = QLabel("⏳ Pending\n17")
        self.pending_kpis_label.setStyleSheet(self.get_card_style("#f39c12"))
        summary_layout.addWidget(self.pending_kpis_label)

        layout.addLayout(summary_layout)

        # KPI Results Table
        self.overview_table = QTableWidget()
        self.overview_table.setColumnCount(7)
        self.overview_table.setHorizontalHeaderLabels([
            "Function", "KPI Name", "Value", "Target", "Status", "Date", "Notes"
        ])
        self.overview_table.horizontalHeader().setStretchLastSection(True)
        self.overview_table.setAlternatingRowColors(True)
        self.overview_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.overview_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #bdc3c7;
                font-size: 10pt;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                font-weight: bold;
                border: none;
            }
        """)

        layout.addWidget(self.overview_table)

        widget.setLayout(layout)
        return widget

    def create_export_tab(self):
        """Create export tab"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Export options
        export_group = QGroupBox("📤 Export KPI Reports")
        export_group.setStyleSheet("""
            QGroupBox {
                font-size: 12pt;
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding: 15px;
            }
        """)
        export_layout = QVBoxLayout()

        info_label = QLabel("Export your KPI data to professional PDF or Excel reports for presentation and archival.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e8f4f8; padding: 10px; border-radius: 5px;")
        export_layout.addWidget(info_label)

        export_layout.addSpacing(20)

        # Export buttons
        export_btn_layout = QHBoxLayout()

        pdf_btn = QPushButton("📄 Export to PDF")
        pdf_btn.clicked.connect(self.export_to_pdf)
        pdf_btn.setStyleSheet(self.get_button_style("#e74c3c"))
        pdf_btn.setMinimumHeight(80)
        pdf_btn.setMinimumWidth(200)
        export_btn_layout.addWidget(pdf_btn)

        excel_btn = QPushButton("📊 Export to Excel")
        excel_btn.clicked.connect(self.export_to_excel)
        excel_btn.setStyleSheet(self.get_button_style("#27ae60"))
        excel_btn.setMinimumHeight(80)
        excel_btn.setMinimumWidth(200)
        export_btn_layout.addWidget(excel_btn)

        export_layout.addLayout(export_btn_layout)
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)

        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def get_card_style(self, color):
        """Get style for summary cards"""
        return f"""
            QLabel {{
                background-color: {color};
                color: white;
                padding: 20px;
                border-radius: 10px;
                font-size: 16pt;
                font-weight: bold;
                text-align: center;
            }}
        """

    def populate_periods(self):
        """Populate period dropdown with quarterly and monthly periods"""
        self.period_combo.clear()
        current = datetime.now()

        # Add separator header for quarterly periods
        self.period_combo.addItem("═══ QUARTERLY PERIODS ═══", None)
        self.period_combo.model().item(0).setEnabled(False)
        self.period_combo.model().item(0).setBackground(QColor("#34495e"))
        self.period_combo.model().item(0).setForeground(QColor("#ffffff"))

        # Add quarterly periods (current year and last 2 years)
        for year in [current.year, current.year - 1, current.year - 2]:
            for quarter in [4, 3, 2, 1]:
                period = f"{year}-Q{quarter}"
                display = f"Q{quarter} {year} (Quarterly)"
                self.period_combo.addItem(display, period)

        # Add separator header for monthly periods
        self.period_combo.addItem("═══ MONTHLY PERIODS ═══", None)
        separator_idx = self.period_combo.count() - 1
        self.period_combo.model().item(separator_idx).setEnabled(False)
        self.period_combo.model().item(separator_idx).setBackground(QColor("#34495e"))
        self.period_combo.model().item(separator_idx).setForeground(QColor("#ffffff"))

        # Add monthly periods (last 12 months)
        for i in range(12):
            month = current.month - i
            year = current.year

            while month < 1:
                month += 12
                year -= 1

            period = f"{year}-{month:02d}"
            display = datetime(year, month, 1).strftime("%B %Y")
            self.period_combo.addItem(display, period)

    def on_period_changed(self, text):
        """Handle period selection change"""
        self.current_period = self.period_combo.currentData()

        # Skip if separator was selected
        if self.current_period is None:
            return

        # Detect if quarterly period
        self.is_quarterly_period = '-Q' in str(self.current_period)

        # Show/hide quarterly report button based on period type
        self.generate_quarterly_btn.setVisible(self.is_quarterly_period)

        # Refresh dashboard
        self.refresh_dashboard()

        # For monthly periods, reload current KPI data if one is selected
        if not self.is_quarterly_period and self.kpi_selector.currentData():
            self.on_kpi_selected(self.kpi_selector.currentText())

    def populate_kpi_selector(self):
        """Populate KPI selector with ALL 17 KPIs"""
        self.kpi_selector.clear()
        self.kpi_selector.addItem("-- Select a KPI --", None)

        all_kpis = self.kpi_manager.get_kpis_needing_manual_data()
        for kpi_name in all_kpis:
            self.kpi_selector.addItem(kpi_name, kpi_name)

    def on_kpi_selected(self, text):
        """Handle KPI selection for data input"""
        # Clear previous form
        while self.input_form_layout.count():
            child = self.input_form_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Clear KPI info
        while self.kpi_info_layout.count():
            child = self.kpi_info_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            return

        # Get required fields
        fields = self.kpi_manager.get_required_fields_for_kpi(kpi_name)

        # Get KPI definition
        kpi_def = self.kpi_manager.get_kpi_by_name(kpi_name)

        if kpi_def:
            # Show KPI info
            info_html = f"""
            <p><b>Description:</b> {kpi_def['description']}</p>
            <p><b>Formula:</b> {kpi_def['formula']}</p>
            <p><b>Target:</b> <span style='color: #27ae60; font-weight: bold;'>{kpi_def['acceptance_criteria']}</span></p>
            <p><b>Frequency:</b> {kpi_def['frequency']}</p>
            """
            info_label = QLabel(info_html)
            info_label.setWordWrap(True)
            info_label.setStyleSheet("padding: 5px; font-size: 10pt;")
            self.kpi_info_layout.addWidget(info_label)

        # Load existing data if any
        existing_data = self.kpi_manager.get_manual_data(kpi_name, self.current_period)
        existing_dict = {row['data_field']: row['data_value'] or row['data_text'] for row in existing_data}

        # Create input fields
        inputs_group = QGroupBox(f"📊 Data Input for: {kpi_name}")
        inputs_group.setStyleSheet("""
            QGroupBox {
                font-size: 11pt;
                font-weight: bold;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding: 15px;
                background-color: #f8f9fa;
            }
        """)
        inputs_layout = QFormLayout()
        inputs_layout.setSpacing(15)
        inputs_layout.setLabelAlignment(Qt.AlignRight)

        self.input_fields = {}

        for field_info in fields:
            field_name = field_info['field']
            label = field_info['label']
            field_type = field_info['type']

            label_widget = QLabel(label + ":")
            label_widget.setStyleSheet("font-weight: bold; color: #2c3e50;")

            if field_type == 'number':
                input_widget = QDoubleSpinBox()
                input_widget.setRange(0, 999999)
                input_widget.setDecimals(2)
                input_widget.setMinimumWidth(200)
                input_widget.setMinimumHeight(35)
                input_widget.setStyleSheet("""
                    QDoubleSpinBox {
                        padding: 5px;
                        border: 2px solid #bdc3c7;
                        border-radius: 5px;
                        font-size: 11pt;
                    }
                    QDoubleSpinBox:focus {
                        border: 2px solid #3498db;
                    }
                """)
                # Load existing value
                if field_name in existing_dict and existing_dict[field_name] is not None:
                    input_widget.setValue(float(existing_dict[field_name]))
            else:  # text
                input_widget = QTextEdit()
                input_widget.setMaximumHeight(100)
                input_widget.setStyleSheet("""
                    QTextEdit {
                        padding: 5px;
                        border: 2px solid #bdc3c7;
                        border-radius: 5px;
                        font-size: 10pt;
                    }
                    QTextEdit:focus {
                        border: 2px solid #3498db;
                    }
                """)
                # Load existing value
                if field_name in existing_dict:
                    input_widget.setText(str(existing_dict[field_name]))

            self.input_fields[field_name] = input_widget
            inputs_layout.addRow(label_widget, input_widget)

        inputs_group.setLayout(inputs_layout)
        self.input_form_layout.addWidget(inputs_group)
        self.input_form_layout.addStretch()

    def save_manual_data(self):
        """Save manual data inputs"""
        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            QMessageBox.warning(self, "No KPI Selected", "Please select a KPI first.")
            return

        try:
            # Save each field
            for field_name, widget in self.input_fields.items():
                if isinstance(widget, QDoubleSpinBox):
                    value = widget.value()
                    self.kpi_manager.save_manual_data(
                        kpi_name=kpi_name,
                        measurement_period=self.current_period,
                        data_field=field_name,
                        data_value=value,
                        entered_by=self.current_user
                    )
                else:  # QTextEdit
                    text = widget.toPlainText()
                    self.kpi_manager.save_manual_data(
                        kpi_name=kpi_name,
                        measurement_period=self.current_period,
                        data_field=field_name,
                        data_value=None,
                        data_text=text,
                        entered_by=self.current_user
                    )

            QMessageBox.information(self, "✓ Success", f"Data saved successfully for {kpi_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save data: {str(e)}\n\n{traceback.format_exc()}")

    def calculate_and_chart(self):
        """Calculate the selected KPI and generate visualization"""
        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            QMessageBox.warning(self, "No KPI Selected", "Please select a KPI first.")
            return

        try:
            # Calculate KPI
            result = self.kpi_manager.calculate_manual_kpi(kpi_name, self.current_period, self.current_user)

            if 'error' in result:
                QMessageBox.warning(self, "Cannot Calculate", result['error'])
                return

            # Generate chart
            self.generate_chart(kpi_name, result)

            # Show success message
            msg = f"✓ KPI Calculated: {kpi_name}\n\n"
            if result.get('value') is not None:
                msg += f"Value: {result['value']:.2f}\n"
            if result.get('text'):
                msg += f"Result: {result['text']}\n"
            if result.get('meets_criteria') is not None:
                status = "✓ PASS" if result['meets_criteria'] else "✗ FAIL"
                msg += f"Status: {status}"

            QMessageBox.information(self, "Calculation Complete", msg)
            self.refresh_dashboard()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to calculate KPI: {str(e)}\n\n{traceback.format_exc()}")

    def generate_chart(self, kpi_name, result):
        """Generate appropriate chart for the KPI"""
        # Clear previous chart
        if self.chart_canvas:
            self.chart_container_layout.removeWidget(self.chart_canvas)
            self.chart_canvas.deleteLater()
            self.chart_canvas = None

        # Create figure
        fig = Figure(figsize=(6, 5), dpi=100)
        fig.patch.set_facecolor('white')

        # Determine chart type based on KPI
        if kpi_name in ['FR1', 'Near Miss', 'Non Conformances raised', 'WO opened vs WO closed',
                        'WO Backlog', 'WO age profile']:
            # Bar chart for count-based KPIs
            self.create_bar_chart(fig, kpi_name, result)
        elif kpi_name in ['Preventive Maintenance Adherence', 'TTR (Time to Repair) Adherence',
                          'Technical Availability Adherence', 'Purchaser satisfaction',
                          'Purchaser Satisfaction Survey', 'Purchaser Monthly process Confirmation',
                          'Non Conformances closed']:
            # Pie/Donut chart for percentage KPIs
            self.create_percentage_chart(fig, kpi_name, result)
        elif kpi_name == 'Top Breakdown':
            # Text display for narrative KPIs
            self.create_text_display(fig, kpi_name, result)
        else:
            # Default: gauge/indicator chart
            self.create_indicator_chart(fig, kpi_name, result)

        # Add canvas to layout
        self.chart_canvas = FigureCanvas(fig)
        self.chart_container_layout.addWidget(self.chart_canvas)
        self.chart_canvas.draw()

    def create_bar_chart(self, fig, kpi_name, result):
        """Create bar chart for count-based KPIs"""
        ax = fig.add_subplot(111)

        # Get data from result text
        value = result.get('value', 0)
        text = result.get('text', '')
        meets_criteria = result.get('meets_criteria')

        if kpi_name == 'WO opened vs WO closed':
            # Parse the text to get opened, closed, currently open
            import re
            opened = int(re.search(r'Opened: (\d+)', text).group(1)) if 'Opened:' in text else 0
            closed = int(re.search(r'Closed: (\d+)', text).group(1)) if 'Closed:' in text else 0
            currently_open = int(re.search(r'Currently Open: (\d+)', text).group(1)) if 'Currently Open:' in text else 0

            categories = ['Opened', 'Closed', 'Currently Open']
            values = [opened, closed, currently_open]
            colors = ['#3498db', '#27ae60', '#e74c3c' if currently_open > 40 else '#f39c12']

            bars = ax.bar(categories, values, color=colors, edgecolor='black', linewidth=1.5)
            ax.set_ylabel('Count', fontweight='bold')
            ax.set_title(f'{kpi_name}\n{self.current_period}', fontweight='bold', fontsize=12)

            # Add value labels on bars
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}',
                       ha='center', va='bottom', fontweight='bold')
        else:
            # Simple single bar chart
            color = '#27ae60' if meets_criteria else '#e74c3c'
            bar = ax.bar([kpi_name.split()[0]], [value], color=color, edgecolor='black', linewidth=1.5, width=0.5)
            ax.set_ylabel('Value', fontweight='bold')
            ax.set_title(f'{kpi_name}\n{self.current_period}', fontweight='bold', fontsize=12)

            # Add value label
            height = bar[0].get_height()
            ax.text(bar[0].get_x() + bar[0].get_width()/2., height,
                   f'{value:.1f}',
                   ha='center', va='bottom', fontweight='bold')

        ax.grid(axis='y', alpha=0.3, linestyle='--')
        fig.tight_layout()

    def create_percentage_chart(self, fig, kpi_name, result):
        """Create donut chart for percentage KPIs"""
        ax = fig.add_subplot(111)

        value = result.get('value', 0)
        meets_criteria = result.get('meets_criteria')

        # Create donut chart
        remaining = max(0, 100 - value)
        values = [value, remaining]
        colors = ['#27ae60' if meets_criteria else '#e74c3c', '#ecf0f1']
        labels = [f'{value:.1f}%', f'{remaining:.1f}%']

        wedges, texts, autotexts = ax.pie(values, labels=labels, colors=colors, autopct='',
                                            startangle=90, counterclock=False,
                                            wedgeprops=dict(width=0.5, edgecolor='white', linewidth=2))

        # Add center text
        ax.text(0, 0, f'{value:.1f}%', ha='center', va='center', fontsize=20, fontweight='bold')

        status = "✓ PASS" if meets_criteria else "✗ FAIL" if meets_criteria is not None else "N/A"
        ax.set_title(f'{kpi_name}\n{self.current_period}\nStatus: {status}',
                    fontweight='bold', fontsize=11)

        fig.tight_layout()

    def create_indicator_chart(self, fig, kpi_name, result):
        """Create indicator/gauge style chart"""
        ax = fig.add_subplot(111)
        ax.axis('off')

        value = result.get('value', 0)
        text = result.get('text', 'N/A')
        meets_criteria = result.get('meets_criteria')

        # Large value display
        color = '#27ae60' if meets_criteria else '#e74c3c' if meets_criteria is not None else '#95a5a6'

        ax.text(0.5, 0.6, f'{value:.2f}' if value is not None else 'N/A',
               ha='center', va='center', fontsize=36, fontweight='bold',
               color=color, transform=ax.transAxes)

        ax.text(0.5, 0.4, text,
               ha='center', va='center', fontsize=10,
               transform=ax.transAxes, wrap=True)

        status = "✓ PASS" if meets_criteria else "✗ FAIL" if meets_criteria is not None else "N/A"
        status_color = '#27ae60' if meets_criteria else '#e74c3c' if meets_criteria is not None else '#95a5a6'

        ax.text(0.5, 0.25, f'Status: {status}',
               ha='center', va='center', fontsize=14, fontweight='bold',
               color=status_color, transform=ax.transAxes)

        ax.text(0.5, 0.85, f'{kpi_name}\n{self.current_period}',
               ha='center', va='center', fontsize=11, fontweight='bold',
               transform=ax.transAxes)

        fig.tight_layout()

    def create_text_display(self, fig, kpi_name, result):
        """Create text display for narrative KPIs"""
        ax = fig.add_subplot(111)
        ax.axis('off')

        text = result.get('text', 'No data entered')

        ax.text(0.5, 0.7, kpi_name,
               ha='center', va='center', fontsize=14, fontweight='bold',
               transform=ax.transAxes)

        ax.text(0.5, 0.4, text,
               ha='center', va='center', fontsize=10,
               transform=ax.transAxes, wrap=True)

        ax.text(0.5, 0.1, self.current_period,
               ha='center', va='center', fontsize=10, style='italic',
               transform=ax.transAxes)

        fig.tight_layout()

    def refresh_dashboard(self):
        """Refresh the overview dashboard"""
        try:
            # Get results based on period type
            if self.is_quarterly_period:
                # Parse quarterly period (e.g., "2025-Q1")
                year, quarter = self.current_period.split('-Q')
                year = int(year)
                quarter = int(quarter)

                # Check if quarterly data exists in database
                results = self.quarterly_calculator.get_quarterly_kpi_results(year, quarter)

                # If no saved quarterly data, calculate it on the fly
                if not results:
                    quarterly_kpis = self.quarterly_calculator.calculate_all_quarterly_kpis(year, quarter)
                    # Convert to result format
                    results = [kpi for kpi in quarterly_kpis if 'error' not in kpi and kpi.get('value') is not None]
            else:
                # Monthly period - use existing logic
                results = self.kpi_manager.get_kpi_results(self.current_period)

            # Update summary cards
            total = len(results)
            passing = sum(1 for r in results if r.get('meets_criteria') is True)
            failing = sum(1 for r in results if r.get('meets_criteria') is False)
            pending = 17 - total

            period_type = "Quarterly" if self.is_quarterly_period else "Monthly"
            self.total_kpis_label.setText(f"Total KPIs\n{total}/17\n({period_type})")
            self.passing_kpis_label.setText(f"✓ Passing\n{passing}")
            self.failing_kpis_label.setText(f"✗ Failing\n{failing}")
            self.pending_kpis_label.setText(f"⏳ Pending\n{pending}")

            # Update table
            self.overview_table.setRowCount(len(results))

            for row, result in enumerate(results):
                # Function
                self.overview_table.setItem(row, 0, QTableWidgetItem(result.get('function_code', '')))

                # KPI Name
                self.overview_table.setItem(row, 1, QTableWidgetItem(result.get('kpi_name', '')))

                # Value
                value_text = result.get('calculated_text') or (
                    f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                )
                self.overview_table.setItem(row, 2, QTableWidgetItem(value_text))

                # Target
                self.overview_table.setItem(row, 3, QTableWidgetItem(result.get('acceptance_criteria', '')))

                # Status
                if result.get('meets_criteria') is True:
                    status_item = QTableWidgetItem("✓ PASS")
                    status_item.setBackground(QColor("#d5f4e6"))
                    status_item.setForeground(QColor("#27ae60"))
                elif result.get('meets_criteria') is False:
                    status_item = QTableWidgetItem("✗ FAIL")
                    status_item.setBackground(QColor("#fadbd8"))
                    status_item.setForeground(QColor("#e74c3c"))
                else:
                    status_item = QTableWidgetItem("N/A")
                    status_item.setBackground(QColor("#f9e79f"))

                status_item.setFont(QFont("Arial", 10, QFont.Bold))
                self.overview_table.setItem(row, 4, status_item)

                # Date
                calc_date = result.get('calculation_date')
                date_str = calc_date.strftime('%Y-%m-%d %H:%M') if calc_date else ''
                self.overview_table.setItem(row, 5, QTableWidgetItem(date_str))

                # Notes
                self.overview_table.setItem(row, 6, QTableWidgetItem(result.get('notes', '')))

            self.overview_table.resizeColumnsToContents()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh dashboard: {str(e)}\n\n{traceback.format_exc()}")

    def export_to_pdf(self):
        """Export KPI data to PDF"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export to PDF", f"KPI_Report_{self.current_period}.pdf", "PDF Files (*.pdf)"
        )

        if not file_name:
            return

        try:
            results = self.kpi_manager.get_kpi_results(self.current_period)

            # Create PDF
            doc = SimpleDocTemplate(file_name, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("KPI Performance Report 2025", title_style))

            # Period info
            period_text = f"Measurement Period: {self.current_period}"
            elements.append(Paragraph(period_text, styles['Normal']))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            elements.append(Paragraph(f"Generated by: {self.current_user}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Summary
            total = len(results)
            passing = sum(1 for r in results if r.get('meets_criteria') is True)
            failing = sum(1 for r in results if r.get('meets_criteria') is False)

            summary_data = [
                ['Total KPIs', 'Passing', 'Failing', 'Pending'],
                [str(total), str(passing), str(failing), str(17 - total)]
            ]

            summary_table = Table(summary_data, colWidths=[2*inch]*4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 30))

            # KPI Details
            if results:
                elements.append(Paragraph("KPI Details", styles['Heading2']))
                elements.append(Spacer(1, 10))

                data = [['Function', 'KPI', 'Value', 'Target', 'Status']]

                for result in results:
                    value_text = result.get('calculated_text') or (
                        f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                    )

                    if result.get('meets_criteria') is True:
                        status = "PASS"
                    elif result.get('meets_criteria') is False:
                        status = "FAIL"
                    else:
                        status = "N/A"

                    data.append([
                        result.get('function_code', ''),
                        Paragraph(result.get('kpi_name', ''), styles['Normal']),
                        Paragraph(str(value_text), styles['Normal']),
                        Paragraph(result.get('acceptance_criteria', ''), styles['Normal']),
                        status
                    ])

                kpi_table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1.8*inch, 1.8*inch, 0.8*inch])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(kpi_table)

            # Build PDF
            doc.build(elements)

            QMessageBox.information(self, "Success", f"PDF exported successfully to:\n{file_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF: {str(e)}\n\n{traceback.format_exc()}")

    def export_to_excel(self):
        """Export KPI data to Excel"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", f"KPI_Report_{self.current_period}.xlsx", "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        try:
            results = self.kpi_manager.get_kpi_results(self.current_period)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "KPI Report"

            # Styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = 'KPI Performance Report 2025'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Period info
            ws['A2'] = f'Period: {self.current_period}'
            ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A4'] = f'Generated by: {self.current_user}'

            # Headers
            row = 6
            headers = ['Function', 'KPI Name', 'Value', 'Target', 'Status', 'Calculated Date', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Data
            for result in results:
                row += 1

                value_text = result.get('calculated_text') or (
                    f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                )

                if result.get('meets_criteria') is True:
                    status = "PASS"
                    status_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                elif result.get('meets_criteria') is False:
                    status = "FAIL"
                    status_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                else:
                    status = "N/A"
                    status_fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")

                calc_date = result.get('calculation_date')
                date_str = calc_date.strftime('%Y-%m-%d %H:%M') if calc_date else ''

                data = [
                    result.get('function_code', ''),
                    result.get('kpi_name', ''),
                    value_text,
                    result.get('acceptance_criteria', ''),
                    status,
                    date_str,
                    result.get('notes', '')
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 5:  # Status column
                        cell.fill = status_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 40

            # Save
            wb.save(file_name)

            QMessageBox.information(self, "Success", f"Excel file exported successfully to:\n{file_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel: {str(e)}\n\n{traceback.format_exc()}")

    def generate_quarterly_report(self):
        """Generate and save quarterly KPI report"""
        if not self.is_quarterly_period:
            QMessageBox.warning(self, "Not a Quarterly Period", "Please select a quarterly period (Q1-Q4) first.")
            return

        try:
            # Parse quarterly period
            year, quarter = self.current_period.split('-Q')
            year = int(year)
            quarter = int(quarter)

            # Show progress dialog
            progress = QProgressDialog("Generating quarterly report...", "Cancel", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(10)

            # Generate quarterly report
            report = self.quarterly_calculator.generate_quarterly_report(
                year=year,
                quarter=quarter,
                save_to_db=True,
                calculated_by=self.current_user
            )

            progress.setValue(80)

            # Refresh dashboard to show new data
            self.refresh_dashboard()

            progress.setValue(100)
            progress.close()

            # Show success message with summary
            stats = report['statistics']
            quarter_info = report['quarter_info']

            msg = f"✓ Quarterly Report Generated Successfully!\n\n"
            msg += f"Quarter: {quarter_info['label']}\n"
            msg += f"Period: {quarter_info['start_date']} to {quarter_info['end_date']}\n\n"
            msg += f"Summary:\n"
            msg += f"  Total KPIs: {stats['total_kpis']}\n"
            msg += f"  With Data: {stats['kpis_with_data']}\n"
            msg += f"  Passing: {stats['kpis_passing']} (✓)\n"
            msg += f"  Failing: {stats['kpis_failing']} (✗)\n"
            msg += f"  Pending: {stats['kpis_pending']}\n\n"

            if stats['kpis_with_data'] > 0:
                pass_rate = (stats['kpis_passing'] / stats['kpis_with_data']) * 100
                msg += f"Pass Rate: {pass_rate:.1f}%\n\n"

            msg += f"The quarterly results have been saved to the database.\n"
            msg += f"You can now export them to PDF or Excel."

            QMessageBox.information(self, "Quarterly Report Generated", msg)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate quarterly report: {str(e)}\n\n{traceback.format_exc()}")
